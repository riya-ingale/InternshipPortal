from flask import Flask, render_template, url_for, request, redirect, session, flash, send_file
from flask import make_response, session, g
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from io import BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
import os
from datetime import datetime
import pdfkit
import flask_excel as excel
from openpyxl import Workbook, load_workbook
import requests
# from office365.runtime.auth.authentication_context import AuthenticationContext
# from office365.sharepoint.client_context import ClientContext
# from office365.sharepoint.files.file import File

app = Flask(__name__)

app.config['SECRET_KEY'] = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///InternshipPortal_Database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

excel.init_excel(app)

db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "You need to Login first"


@login_manager.user_loader
def load_user(user_id):
    return Users.query.get(int(user_id))


class Users(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fullname = db.Column(db.String(150))
    rollno = db.Column(db.String(15), unique=True)
    password = db.Column(db.String(80))
    mobileno = db.Column(db.String(15))
    email = db.Column(db.String(50))
    dept = db.Column(db.String(15))
    div = db.Column(db.String(15))
    year = db.Column(db.String(15))
    linkedIn_username = db.Column(db.Text)
    role = db.Column(db.Text)


class Internships(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer)
    companyname = db.Column(db.Text)
    position = db.Column(db.Text)
    domain = db.Column(db.Text)
    source = db.Column(db.Text)
    skills_acquired = db.Column(db.Text)
    companyrepresentative_name = db.Column(db.Text)
    companyrepresentative_contact = db.Column(db.Text)
    startdate = db.Column(db.Date)
    enddate = db.Column(db.Date)
    offerletter = db.Column(db.LargeBinary)
    offerletter_filename = db.Column(db.Text)
    completioncert = db.Column(db.LargeBinary)
    completioncert_filename = db.Column(db.Text)
    feedback = db.Column(db.Text)
    workenv = db.Column(db.Integer)
    satisfied = db.Column(db.Text)
    recommendation = db.Column(db.Text)
    typeofinternship = db.Column(db.Text)


@app.route('/')
def index():
    return render_template("index.html")


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == "POST":
        fullname = request.form.get('fullname')
        rollno = request.form['rollno']
        email = request.form['email']
        mobileno = request.form['mobileno']
        dept = request.form['dept']
        div = request.form['div']
        year = request.form['year']
        password = request.form['password']
        hashed_password = generate_password_hash(password, method="sha256")
        cpassword = request.form['cpassword']

        user = Users.query.filter_by(rollno=rollno).first()
        if user:
            flash("Roll number Already Exists")
            return redirect("/signup")
        user = Users.query.filter_by(email=email).first()
        if user:
            flash("Email Already Registered")
            return redirect("/login")

        if(password == cpassword):
            new_user = Users(fullname=fullname, rollno=rollno, email=email, mobileno=mobileno,
                             dept=dept, div=div, password=hashed_password, year=year)
            db.session.add(new_user)
            db.session.commit()

            flash("Sucessfully Registered!", "success")
            return redirect('/login')
        else:
            flash("Passwords don't match", "danger")
            return redirect("/signup")

    return render_template("login.html")


@app.route('/login', methods=['POST', 'GET'])
def login():
    print("Session next", session.get('next'))
    if request.method == "POST":
        logout_user()
        rollno = request.form.get('rollno')
        password = request.form.get('password')

        user = Users.query.filter_by(rollno=rollno).first()
        if not user:
            flash("No such User found, Try Signing Up First", "warning")
            return redirect("/signup")
        if user:
            if check_password_hash(user.password, password):
                login_user(user)
                if session.get("next"):
                    return redirect("/"+session.get("next"))
                return redirect("/")
            else:
                flash("Incorrect password", "danger")
                return redirect("login")
    return render_template("login.html")


@app.route('/logout')
def logout():
    logout_user()
    return redirect('/')


@app.route('/profile/<int:user_id>', methods=['GET', 'POST'])
@app.route('/profile/', methods=['GET', 'POST'])
@login_required
def profile(user_id):
    user = Users.query.filter_by(id=user_id).first()
    internships = Internships.query.filter_by(user_id=user_id).all()
    return render_template("profile.html", current_user=current_user, internships=internships, user=user)


@app.route('/otherprofile/<int:user_id>', methods=['GET', 'POST'])
def otherprofile(user_id):
    user = Users.query.filter_by(id=user_id).first()
    internships = Internships.query.filter_by(user_id=user_id).all()
    return render_template("otherprofile.html", current_user=current_user, internships=internships, user=user)


@app.route('/search', methods=['GET', 'POST'])
def search():
    if not current_user.is_authenticated:
        return redirect('/admin/login')
    else:
        if current_user.role == "admin":
            allinternships = []
            allstudents = []
            student_data = []
            internship_data = []
            if request.method == "GET":
                allstudents = Users.query.filter(Users.id != 2).all()
                allinternships = Internships.query.all()
                return render_template("search.html", students=allstudents, internships=allinternships, s=True)

            if request.method == 'POST':
                searchname = request.form.get('searchname')
                dept = request.form.get('dept')
                div = request.form.get('div')
                year = request.form.get('year')
                domain = request.form.get('domain')
                satisfied = request.form.get('satisfied')
                startdate = request.form.get('startdate')
                rollno = request.form.get('rollno')
                if startdate:
                    startdate = datetime.strptime(startdate, '%Y-%m-%d')
                enddate = request.form.get('enddate')
                if enddate:
                    enddate = datetime.strptime(enddate, '%Y-%m-%d')

                search = "{0}".format(searchname)
                search = search+'%'

                if startdate and enddate and searchname:
                    allinternships = Internships.query.filter(
                        or_(Internships.companyname.like(search), Internships.domain.like(search)), Internships.startdate > startdate and Internships.enddate < enddate).all()

                elif startdate and searchname and not enddate:
                    allinternships = Internships.query.filter(
                        or_(Internships.companyname.like(search), Internships.domain.like(search)), Internships.startdate > startdate).all()
                elif enddate and searchname and not startdate:
                    allinternships = Internships.query.filter(
                        or_(Internships.companyname.like(search), Internships.domain.like(search)), Internships.endate < enddate).all()
                elif startdate and enddate and not searchname:
                    allinternships = Internships.query.filter(
                        Internships.startdate > startdate and Internships.enddate < enddate).all()
                elif searchname and not enddate and not startdate:
                    allinternships = Internships.query.filter(
                        or_(Internships.companyname.like(search), Internships.domain.like(search))).all()
                elif startdate and not searchname and not enddate:
                    allinternships = Internships.query.filter(
                        Internships.startdate > startdate).all()
                elif enddate and not startdate and not searchname:
                    allinternships = Internships.query.filter(
                        Internships.enddate < enddate).all()
                else:
                    pass

                if not allinternships:
                    if domain and satisfied:
                        allinternships = Internships.query.filter_by(
                            domain=domain, satisfied=satisfied).all()
                    elif domain and not satisfied:
                        allinternships = Internships.query.filter_by(
                            domain=domain).all()
                    elif satisfied and not domain:
                        allinternships = Internships.query.filter_by(
                            satisfied=satisfied).all()
                    else:
                        pass
                else:
                    if domain and satisfied:
                        for internship in allinternships:
                            if internship.domain == domain and internship.satisfied == satisfied:
                                pass
                            else:
                                allinternships.remove(internship)
                    elif domain and not satisfied:
                        for internship in allinternships:
                            if internship.domain == domain:
                                pass
                            else:
                                allinternships.remove(internship)
                    elif satisfied and not domain:
                        for internship in allinternships:
                            if internship.satisfied == satisfied:
                                pass
                            else:
                                allinternships.remove(internship)
                    else:
                        pass

                if dept and div and year:
                    allstudents = Users.query.filter_by(
                        dept=dept, div=div, year=year).all()
                elif dept and div and not year:
                    allstudents = Users.query.filter_by(
                        dept=dept, div=div).all()
                elif div and year and not dept:
                    allstudents = Users.query.filter_by(
                        div=div, year=year).all()
                elif dept and year and not div:
                    allstudents = Users.query.filter_by(
                        dept=dept, year=year).all()
                elif dept and not div and not year:
                    allstudents = Users.query.filter_by(dept=dept).all()
                elif div and not dept and not year:
                    allstudents = Users.query.filter_by(div=div).all()
                elif year and not div and not dept:
                    allstudents = Users.query.filter_by(year=year).all()
                else:
                    pass

                if not allstudents:
                    if rollno:
                        allstudents = Users.query.filter_by(
                            rollno=rollno).all()
                    else:
                        pass
                else:
                    if rollno:
                        for student in allstudents:
                            if student.rollno == rollno:
                                pass
                            else:
                                allstudents.remove(student)
                    else:
                        pass

                if allinternships and not allstudents:
                    for internship in allinternships:
                        student = Users.query.filter_by(
                            id=internship.user_id).first()
                        student_data.append(student)
                    return render_template("search.html", students=student_data, internships=allinternships, s=True)
                elif allstudents and not allinternships:
                    for student in allstudents:
                        student = Internships.query.filter_by(
                            user_id=student.id).first()
                        internship_data.append(student)
                    return render_template("search.html", students=allstudents, internships=internship_data, s=True)
                elif allinternships and allstudents:
                    return render_template("search.html", students=allstudents, internships=allinternships, s=True)
                else:
                    return render_template("search.html", s=False)
        else:
            return "Page for Admin User Only"


@app.route('/newinternship', methods=['GET', 'POST'])
@login_required
def newinternship():
    if request.method == "POST":
        companyname = request.form.get('companyname')
        domain = request.form.get('domain')
        source = request.form.get('source')
        position = request.form.get('position')
        skills_acquired = request.form.get('skills_acquired')
        companyrepresentative_name = request.form.get(
            'companyrepresentative_name')
        companyrepresentative_contact = request.form.get(
            'companyrepresentative_contact')
        startdate = request.form.get('startdate')
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = request.form.get('enddate')
        enddate = datetime.strptime(enddate, '%Y-%m-%d')
        offerletter = request.files['offerletter']
        completioncert = request.files['completioncert']
        if offerletter:
            if len(offerletter.filename) > 0:
                offerletter_filename = offerletter.filename
                offerletter = offerletter.read()
        if completioncert:
            if len(completioncert.filename) > 0:
                completioncert_filename = completioncert.filename
                completioncert = completioncert.read()
        feedback = request.form.get('feedback')
        workenv = request.form.get('workenv')
        satisfied = request.form.get('satisfied')
        recommendation = request.form.get('recommendation')
        typeofinternship = request.form.get('type')
        new_internship = Internships(user_id=current_user.id, companyname=companyname, domain=domain, companyrepresentative_name=companyrepresentative_name, companyrepresentative_contact=companyrepresentative_contact, source=source, position=position, skills_acquired=skills_acquired, startdate=startdate, enddate=enddate,
                                     offerletter=offerletter, offerletter_filename=offerletter_filename, completioncert=completioncert, completioncert_filename=completioncert_filename, feedback=feedback, workenv=workenv, satisfied=satisfied, recommendation=recommendation, typeofinternship=typeofinternship)
        db.session.add(new_internship)
        db.session.commit()
        flash("Record Added!")
        return redirect('/newinternship')
    return render_template("newinternship.html")


@app.route('/updateinternship/<int:id>', methods=['GET', 'POST'])
@login_required
def updateinternship(id):
    internship = Internships.query.filter_by(id=id).first()
    print(internship.companyname)
    if request.method == 'POST':
        internship.companyname = request.form.get('companyname')
        internship.position = request.form.get('position')
        internship.domain = request.form.get('domain')
        internship.source = request.form.get('source')
        internship.rating = request.form.get('rating')
        internship.skills_acquired = request.form.get('skills_acquired')
        internship.companyrepresentative_name = request.form.get(
            'companyrepresentative_name')
        internship.companyrepresentative_contact = request.form.get(
            'companyrepresentative_contact')

        startdate = request.form.get('startdate')
        internship.startdate = datetime.strptime(startdate, '%Y-%m-%d')
        enddate = request.form.get('enddate')
        internship.enddate = datetime.strptime(enddate, '%Y-%m-%d')

        offerletter = request.files['offerletter']
        completioncert = request.files['completioncert']
        if len(offerletter.filename) > 0:
            internship.offerletter_filename = offerletter.filename
            internship.offerletter = offerletter.read()
        if len(completioncert.filename) > 0:
            internship.completioncert_filename = completioncert.filename
            internship.completioncert = completioncert.read()

        internship.typeofinternship = request.form.get('type')
        internship.satisfied = request.form.get('satisfied')
        internship.workenv = request.form.get('workenv')
        internship.recommendation = request.form.get('recommendation')
        internship.feedback = request.form.get('feedback')
        db.session.commit()
        return redirect(f'/profile/{current_user.id}')

    return render_template('updateinternship.html', internship=internship)


@app.route('/downloadcompletioncert/<int:internship_id>', methods=['GET', 'POST'])
def downloadcompletioncert(internship_id):
    internship = Internships.query.filter_by(id=internship_id).first()
    if internship.completioncert:
        file_data = internship.completioncert
        return send_file(BytesIO(file_data), attachment_filename=internship.companyname + "Completioncert.pdf", as_attachment=True)
    else:
        flash("No file Exists")
        return redirect(f'/profile/{current_user.id}')


@app.route('/student_record/<int:user_id>', methods=['POST', 'GET'])
def student_record(user_id):
    user = Users.query.get_or_404(user_id)
    internships = Internships.query.filter_by(user_id=user_id).all()
    rendered = render_template(
        'studentrecorddownload.html', student=user, internships=internships)
    pdf = pdfkit.from_string(rendered, False)
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename = student.rollno +_Details.pdf'

    return response


@app.route('/admin/login', methods=['GET', 'POST'])
def adminlogin():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = Users.query.filter_by(role="admin", fullname=username).first()
        if user:
            if check_password_hash(user.password, password):
                login_user(user)
                return redirect('/admin/dashboard')
            else:
                flash("Password Incorrect")
                return redirect('/admin/login')
        else:
            flash("No such admin exists")
            return redirect('/admin/login')
    return render_template('adminlogin.html')


def save_excel(form_excel):
    _, f_ext = os.path.splitext(form_excel.filename)
    excel_fn = "sheet" + f_ext
    excel_path = os.path.join(app.root_path, excel_fn)
    form_excel.save(excel_path)
    return excel_fn


@app.route('/admin/dashboard', methods=['GET', 'POST'])
def admindashboard():
    if current_user.is_authenticated:
        if current_user.role == "admin":
            if request.method == "POST":
                sheet = request.files['Excel']
                data_file = save_excel(sheet)
                # Load the entire workbook.
                wb = load_workbook(data_file, data_only=True)
                # Load one worksheet.
                ws = wb['TE A']
                all_rows = list(ws.rows)

                # Pull information from specific cells.
                for row in all_rows[2:]:
                    rollno = row[0].value
                    fullname = row[1].value
                    # dept = row[2].value
                    # div = row[3].value
                    # mobileno = row[4].value
                    # email = row[5].value
                    # year = row[6].value

                    companyname = row[2].value
                    # position = row[8].value
                    domain = row[3].value
                    # skills_acquired = row[11].value
                    # companyrepresentative_name = row[12].value
                    # companyrepresentative_contact = row[13].value
                    startdate = row[4].value
                    enddate = row[5].value
                    source = row[6].value
                    if row[6].value == "Yes":
                        certificate_url = row[8].value
                        # url = request.args['certificate_url']  # user provides url in query string
                        r = requests.get(certificate_url)

                        # write to a file in the app's instance folder
                        # come up with a better file name
                        with app.open_instance_resource('downloaded_file', 'wb') as f:
                            f.write(r.content)
                    # feedback = row[16].value
                    # workenv = row[17].value
                    # satisfied = row[18].value
                    # recommendation = row[19].value

                    # startdate = datetime.strptime(startdate, '%Y-%m-%d')
                    # enddate = datetime.strptime(enddate, '%Y-%m-%d')
                    # if startdate:
                    #     startdate = datetime.strptime(startdate, '%d-%m-%Y')
                    # if enddate:
                    #     enddate = datetime.strptime(enddate, '%d-%m-%Y')
                    user = Users.query.filter_by(
                        rollno=rollno, fullname=fullname).first()
                    if user:
                        user_id = user.id
                        if companyname:
                            newinternship = Internships(user_id=user_id, companyname=companyname,
                                                        domain=domain, startdate=startdate, enddate=enddate, source=source)
                            db.session.add(newinternship)
                            db.session.commit()
                    else:
                        newstudent = Users(
                            fullname=fullname, rollno=rollno, password='sha256$cBl7wrlwRwy9QHJB$7a873cb0e1cd6cd2070c00147540fc6ba209e9114152385c94e06fb641951076')
                        db.session.add(newstudent)
                        db.session.commit()
                        student = Users.query.filter_by(
                            rollno=rollno, fullname=fullname).first()
                        user_id = student.id
                        if companyname:
                            newinternship = Internships(user_id=user_id, companyname=companyname,
                                                        domain=domain, startdate=startdate, enddate=enddate, source=source)
                            db.session.add(newinternship)
                            db.session.commit()
                flash('Record Added')
                return redirect('/admin/dashboard')
            return render_template('admindashboard.html')
        else:
            return "Page for Admin Users Only"
    else:
        return "Page Not Found"


@app.route('/editprofile/<int:user_id>', methods=['GET', 'POST'])
def editprofile(user_id):
    user = Users.query.get_or_404(user_id)
    internships = Internships.query.filter_by(user_id=user_id).all()
    if request.method == "POST":
        user = Users.query.get_or_404(user_id)
        user.fullname = request.form.get('fullname')
        user.rollno = request.form['rollno']
        user.email = request.form['email']
        user.mobileno = request.form['mobileno']
        user.dept = request.form['dept']
        user.div = request.form['div']
        user.year = request.form['year']
        db.session.commit()
        return redirect(f'/profile/{user_id}')
    return render_template('editprofile.html', user=user, internships=internships)


@app.route('/aboutus')
def aboutus():
    return render_template('aboutus.html')


@app.route('/exportall', methods=['GET', 'POST'])
def exportall():
    students = Users.query.filter(Users.id != 2).all()
    internships = Internships.query.all()
    added = 0
    if students:
        # WorkBook Info
        print("creating workbook")
        wb = Workbook()
        # insert value in the cells
        ws = wb.active
        ws.title = "Students Data"
        headings = ["Fullname", "Rollno", "Mobileno", "Email", "Department", "Division", "Year", "Company Name", "Position", "Domain", "Source", "skills_required",
                    "Company Representative Name", "Company Representative Contact", "Start Date", "End Date", "Feedback", "Work Environment Rating", "Satisfaction", "Would student recommend?", "Completion Certificate"]
        ws.append(headings)
        for student in students:
            record_student = [student.fullname, student.rollno, student.mobileno,
                              student.email, student.dept, student.div, student.year]
            for internship in internships:
                if internship:
                    if student.id == internship.user_id:
                        if internship.startdate:
                            internship.startdate = internship.startdate.strftime(
                                "%d /%m /%y")
                        if internship.enddate:
                            internship.enddate = internship.enddate.strftime(
                                "%d/ %m/ %y")
                        record_internship = [internship.companyname, internship.position, internship.domain, internship.source, internship.skills_acquired,
                                             internship.companyrepresentative_name, internship.companyrepresentative_contact, internship.startdate, internship.enddate, internship.feedback, internship.workenv, internship.satisfied, internship.recommendation, '=HYPERLINK("{}", "{}")'.format(f"http://127.0.0.1:5000/downloadcompletioncert/{internship.id}", "Download Cert")]
                        record = record_student + record_internship
                        ws.append(record)
                        added = 1
                        record_internship = []   
            if added == 0:        
                record = record_student + record_internship
                ws.append(record)
            if added == 1:    
                added = 0
        wb.save(filename='sample_book.xlsx')
        print("Saved Excel")
        return send_file('sample_book.xlsx', as_attachment=True, download_name='sample_book.xlsx')
    else:
        return None


@app.route("/customexport", methods=['GET', 'POST'])
def docustomexport():
    allinternships = []
    allstudents = []
    student_data = []
    internship_data = []
    searchname = request.form.get('searchname')
    dept = request.form.get('dept')
    div = request.form.get('div')
    year = request.form.get('year')
    satisfied = request.form.get('satisfied')
    startdate = request.form.get('startdate')
    domain = request.form.get('domain')
    satisfied = request.form.get('satisfied')
    rollno = request.form.get('rollno')

    print(startdate)
    if startdate:
        startdate = datetime.strptime(startdate, '%Y-%m-%d')
    enddate = request.form.get('enddate')
    print(enddate)
    if enddate:
        enddate = datetime.strptime(enddate, '%Y-%m-%d')

    search = "{0}".format(searchname)
    search = search+'%'
    print(search)

    if startdate and enddate and searchname:
        allinternships = Internships.query.filter(or_(Internships.companyname.like(search), Internships.domain.like(
            search)), Internships.startdate > startdate and Internships.enddate < enddate).all()
    elif startdate and searchname and not enddate:
        allinternships = Internships.query.filter(or_(Internships.companyname.like(
            search), Internships.domain.like(search)), Internships.startdate > startdate).all()
    elif enddate and searchname and not startdate:
        allinternships = Internships.query.filter(or_(Internships.companyname.like(
            search), Internships.domain.like(search)), Internships.endate < enddate).all()
    elif startdate and enddate and not searchname:
        allinternships = Internships.query.filter(
            Internships.startdate > startdate and Internships.enddate < enddate).all()
    elif searchname and not enddate and not startdate:
        allinternships = Internships.query.filter(
            or_(Internships.companyname.like(search), Internships.domain.like(search))).all()
    elif startdate and not searchname and not enddate:
        allinternships = Internships.query.filter(
            Internships.startdate > startdate).all()
    elif enddate and not startdate and not searchname:
        allinternships = Internships.query.filter(
            Internships.enddate < enddate).all()
    else:
        pass

    if not allinternships:
        if domain and satisfied:
            allinternships = Internships.query.filter_by(
                domain=domain, satisfied=satisfied).all()
        elif domain and not satisfied:
            allinternships = Internships.query.filter_by(
                domain=domain).all()
        elif satisfied and not domain:
            allinternships = Internships.query.filter_by(
                satisfied=satisfied).all()
        else:
            pass
    else:
        if domain and satisfied:
            for internship in allinternships:
                if internship.domain == domain and internship.satisfied == satisfied:
                    pass
                else:
                    allinternships.remove(internship)
        elif domain and not satisfied:
            for internship in allinternships:
                if internship.domain == domain:
                    pass
                else:
                    allinternships.remove(internship)
        elif satisfied and not domain:
            for internship in allinternships:
                if internship.satisfied == satisfied:
                    pass
                else:
                    allinternships.remove(internship)
        else:
            pass

    if dept and div and year:
        allstudents = Users.query.filter_by(
            dept=dept, div=div, year=year).all()
    elif dept and div and not year:
        allstudents = Users.query.filter_by(
            dept=dept, div=div).all()
    elif div and year and not dept:
        allstudents = Users.query.filter_by(
            div=div, year=year).all()
    elif dept and year and not div:
        allstudents = Users.query.filter_by(
            dept=dept, year=year).all()
    elif dept and not div and not year:
        allstudents = Users.query.filter_by(dept=dept).all()
    elif div and not dept and not year:
        allstudents = Users.query.filter_by(div=div).all()
    elif year and not div and not dept:
        allstudents = Users.query.filter_by(year=year).all()
    else:
        pass

    if not allstudents:
        if rollno:
            allstudents = Users.query.filter_by(rollno=rollno).all()
        else:
            pass
    else:
        if rollno:
            for student in allstudents:
                if student.rollno == rollno:
                    pass
                else:
                    allstudents.remove(student)
        else:
            pass

    if allinternships and not allstudents:
        for internship in allinternships:
            student = Users.query.filter_by(id=internship.user_id).first()
            student_data.append(student)
        students = student_data
        internships = allinternships
        # return render_template("search.html", students=student_data, internships=allinternships, s = True)
    elif allstudents and not allinternships:
        for student in allstudents:
            student = Internships.query.filter_by(
                user_id=student.id).first()
            internship_data.append(student)
        students = allstudents
        internships = internship_data
        # return render_template("search.html", students=allstudents, internships=internship_data, s = True)
    elif allinternships and allstudents:
        students = allstudents
        internships = allinternships
        # return render_template("search.html", students=allstudents, internships=allinternships, s = True)
    else:
        students = []
        internships = []
        pass
        # return render_template("search.html",s = False)

    added = 0
    if students:
        # WorkBook Info
        print("creating workbook")
        print("Students:", students)
        wb = Workbook()
        # insert value in the cells
        ws = wb.active
        ws.title = "Students Data"
        headings = ["Fullname", "Rollno", "Mobileno", "Email", "Department", "Division", "Year", "Company Name", "Position", "Domain", "Source", "skills_required",
                    "Company Representative Name", "Company Representative Contact", "Start Date", "End Date", "Feedback", "Work Environment Rating", "Satisfaction", "Would student recommend?", "Completition Certificate"]
        ws.append(headings)
        for student in students:
            print("Particular Student:", student)
            record_student = [student.fullname, student.rollno, student.mobileno,
                              student.email, student.dept, student.div, student.year]
            print("Internships:", internships)
            for internship in internships:
                if internship:
                    if student.id == internship.user_id:
                        if internship.startdate:
                            internship.startdate = internship.startdate.strftime(
                                "%d /%m /%y")
                        if internship.enddate:
                            internship.enddate = internship.enddate.strftime(
                                "%d/ %m/ %y")
                        record_internship = [internship.companyname, internship.position, internship.domain, internship.source, internship.skills_acquired,
                                             internship.companyrepresentative_name, internship.companyrepresentative_contact, internship.startdate, internship.enddate, internship.feedback, internship.workenv, internship.satisfied, internship.recommendation, '=HYPERLINK("{}", "{}")'.format(f"http://127.0.0.1:5000/downloadcompletioncert/{internship.id}", "Download Cert")]
                        record = record_student + record_internship
                        ws.append(record)
                        added = 1
                        record_internship = []
            if added == 0:        
                record = record_student + record_internship
                ws.append(record)
            if added == 1:    
                added = 0            
        wb.save(filename='sample_book.xlsx')
        return send_file('sample_book.xlsx', as_attachment=True, download_name='sample_book.xlsx')
    else:
        return redirect('/search')


@app.route('/excelformatdownload')
def excelformatdownload():
    return send_file('Excel Format for Upload to Internship Portal.xlsx', as_attachment=True, download_name='Excel Format for Upload to Internship Portal.xlsx')


if __name__ == '__main__':
    db.create_all()
    app.run(debug=True)
